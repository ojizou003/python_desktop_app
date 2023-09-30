from pathlib import Path
import openpyxl

cd = Path.cwd()
filepath = cd/'Exeloperation\sample.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb['シート1']
# a4 = ws.cell(4,1).value
# c2 = ws.cell(2,3).value
cell_range = ws["a1:c5"]
print(cell_range)
# a4 = ws['a4'].value
# c2 = ws['c2'].value
# print(f"a4 = {a4},c2 = {c2}")
