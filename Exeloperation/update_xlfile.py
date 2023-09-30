from pathlib import Path
import openpyxl

cd = Path.cwd()
filepath = cd/'Exeloperation\\newfile.xlsx'
wb = openpyxl.load_workbook(filepath)
ws = wb['MySheet1']
ws.cell(row=4,column=1,value='A4')
ws.cell(row=3,column=2,value='B3')
wb.save(filepath)