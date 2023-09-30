import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'MySheet1'
ws.cell(row=2,column=1,value='A2')
ws.cell(row=4,column=3,value='C4')
wb.save('Exeloperation\\newfile.xlsx')
